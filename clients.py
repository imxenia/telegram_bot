import openpyxl
from openpyxl import Workbook
import os

EXCEL_FILE = 'clients_data.xlsx'

def create_excel_file_if_not_exists():
    if not os.path.exists(EXCEL_FILE):
        print(f"Файл {EXCEL_FILE} не существует. Создаём новый файл...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"
        ws.append(["Номер", "ID клиента", "Имя", "Фамилия", "Услуга", "Доп.услуга", "Дата", "Время", "Номер телефона"])

        wb.save(EXCEL_FILE)
        print(f"Файл {EXCEL_FILE} создан успешно.")
    else:
        print(f"Файл {EXCEL_FILE} уже существует.")

def save_client_data(user_id, name, surname, service, added_service, day, time, phone):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    client_exists = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == user_id and row[2] == name and row[3] == surname:
            client_exists = True
            client_row = row[0]
            break

    if client_exists:
        for row in ws.iter_rows(min_row=2):
            if row[0].value == client_row:
                row[4].value = service
                row[5].value = added_service
                row[6].value = day
                row[7].value = time
                row[8].value = phone
                print(f"Данные клиента с user_id {user_id} обновлены.")
                break
    else:
        next_id = ws.max_row
        ws.append([next_id, user_id, name, surname, service, added_service, day, time, phone])
        print(f"Новый клиент {name} {surname} (user_id: {user_id}) добавлен.")

    wb.save(EXCEL_FILE)
